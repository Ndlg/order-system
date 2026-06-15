import json
import os
import re
from datetime import datetime, timedelta, timezone
from io import BytesIO
from pathlib import Path
import zipfile


TEST_DB = Path(__file__).resolve().parent / "foundation_test.db"
TEST_STORAGE = Path(__file__).resolve().parent / "foundation_storage"
if TEST_DB.exists():
    TEST_DB.unlink()
if TEST_STORAGE.exists():
    import shutil

    shutil.rmtree(TEST_STORAGE)

os.environ["DATABASE_URL"] = f"sqlite:///{TEST_DB.as_posix()}"
os.environ["AUTO_CREATE_TABLES"] = "true"
os.environ["SECRET_KEY"] = "test-secret"
os.environ["STORAGE_ROOT"] = TEST_STORAGE.as_posix()

from fastapi.testclient import TestClient  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from app.main import app  # noqa: E402
from app.models import Collector  # noqa: E402


def test_auth_and_workspace_scoped_field_definition_flow() -> None:
    with TestClient(app) as client:
        login = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
        assert login.status_code == 200
        token = login.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}", "X-Workspace-Id": "1"}

        me = client.get("/api/v1/auth/me", headers=headers)
        assert me.status_code == 200
        assert me.json()["username"] == "admin"
        assert me.json()["tenant_ids"] == [1]
        assert me.json()["workspaces"][0]["id"] == 1
        assert me.json()["workspaces"][0]["tenant_id"] == 1

        tenants = client.get("/api/v1/tenants", headers=headers)
        assert tenants.status_code == 200
        assert tenants.json()[0]["code"] == "default"

        workspace_update = client.patch(
            "/api/v1/workspaces/1",
            headers=headers,
            json={"tenant_id": 999, "name": "Default workspace renamed"},
        )
        assert workspace_update.status_code == 200
        assert workspace_update.json()["tenant_id"] == 1
        assert workspace_update.json()["name"] == "Default workspace renamed"

        waybill_modes = client.get("/api/v1/waybill-modes", headers=headers)
        assert waybill_modes.status_code == 200
        assert waybill_modes.json()[0]["code"] == "manual_upload"
        assert "tenant_id" not in waybill_modes.json()[0]
        assert "workspace_id" not in waybill_modes.json()[0]

        global_mode = client.post(
            "/api/v1/waybill-modes",
            headers=headers,
            json={
                "name": "Global parser",
                "code": "global_parser",
                "input_format": "json",
                "tenant_id": 999,
                "workspace_id": 999,
                "is_deleted": True,
            },
        )
        assert global_mode.status_code == 201
        assert "tenant_id" not in global_mode.json()
        assert "workspace_id" not in global_mode.json()
        assert global_mode.json()["is_deleted"] is False

        created = client.post(
            "/api/v1/field-definitions",
            headers=headers,
            json={
                "name": "Field A",
                "code": "field_a",
                "data_type": "text",
                "tenant_id": 999,
                "workspace_id": 999,
                "created_by": 999,
                "is_deleted": True,
                "export_enabled": False,
            },
        )
        assert created.status_code == 201
        assert created.json()["tenant_id"] == 1
        assert created.json()["workspace_id"] == 1
        assert created.json()["created_by"] == 1
        assert created.json()["is_deleted"] is False

        records = client.get("/api/v1/field-definitions", headers=headers)
        assert records.status_code == 200
        assert len(records.json()) == 1

        legacy_collector_login = client.post(
            "/api/v1/collector-runtime/login",
            json={
                "username": "admin",
                "password": "admin123",
                "workspace_id": 1,
            },
        )
        assert legacy_collector_login.status_code == 404

        collector_registration = client.post(
            "/api/v1/collector-control/register",
            headers=headers,
            json={
                "collector_id": "local-simulator",
                "collector_name": "Local simulator",
                "source_machine": "dev-machine",
                "client_version": "0.1.0",
            },
        )
        assert collector_registration.status_code == 201
        collector_body = collector_registration.json()
        collector_token = collector_body["collector_token"]
        collector = collector_body["collector"]
        assert collector["tenant_id"] == 1
        assert collector["workspace_id"] == 1
        assert "token_hash" not in collector

        collectors = client.get("/api/v1/collectors", headers=headers)
        assert collectors.status_code == 200
        assert "token_hash" not in collectors.json()[0]

        collector_exe = Path(__file__).resolve().parents[2] / "collector-client" / "dist" / "订单系统采集器.exe"
        created_collector_stub = False
        if not collector_exe.exists():
            collector_exe.parent.mkdir(parents=True, exist_ok=True)
            collector_exe.write_bytes(b"collector exe stub")
            created_collector_stub = True
        try:
            collector_download = client.get("/api/v1/collector-client/download", headers=headers)
            assert collector_download.status_code == 200
            assert collector_download.headers["content-type"] == "application/zip"
            with zipfile.ZipFile(BytesIO(collector_download.content)) as archive:
                names = set(archive.namelist())
                assert "订单系统采集器/订单系统采集器.exe" in names
                assert "订单系统采集器/VERSION.txt" in names
                assert "订单系统采集器/参数说明.txt" in names
                assert not any(name.endswith((".bat", ".vbs", ".json", ".py")) for name in names)
                assert "订单系统采集器/order-system-collector-agent.exe" not in names
                assert "订单系统采集器/order-system-collector-cli.exe" not in names
                assert not any("__pycache__" in name for name in names)
                version_text = archive.read("订单系统采集器/VERSION.txt").decode("utf-8")
                assert "version=single-exe-token-collector-20260614" in version_text
                assert "mode=cli" in version_text
                assert "token-only" in version_text
                guide_text = archive.read("订单系统采集器/参数说明.txt").decode("utf-8")
                assert "订单系统采集器.exe" in guide_text
                assert "业务机不再输入系统账号密码" in guide_text
                assert not re.search(r"\b10\.139\.\d{1,3}\.\d{1,3}\b", guide_text)
                assert "biz-left" not in guide_text
                assert "Windows 机器名" in guide_text
                assert "collector_login" not in guide_text
        finally:
            if created_collector_stub:
                collector_exe.unlink()

        capture_task = client.post(
            "/api/v1/collector-control/start",
            headers=headers,
            json={"name": "Test capture"},
        )
        assert capture_task.status_code == 201
        assert capture_task.json()["status"] == "collecting"

        heartbeat = client.post(
            "/api/v1/collector-runtime/heartbeat",
            headers={"X-Collector-Token": collector_token},
            json={
                "source_machine": "dev-machine",
                "collector_id": "dev-machine",
                "queue_size": 0,
                "adapter_status": {"simulator": {"status": "ready"}},
            },
        )
        assert heartbeat.status_code == 200
        assert heartbeat.json()["collector"]["online_status"] == "online"
        assert heartbeat.json()["collector"]["collector_id"] == "dev-machine"
        assert heartbeat.json()["collector"]["source_machine"] == "dev-machine"
        assert heartbeat.json()["collector"]["status_payload"]["adapter_status"]["simulator"]["status"] == "ready"
        assert heartbeat.json()["tasks"][0]["id"] == capture_task.json()["id"]

        with SessionLocal() as db:
            stale_collector = db.get(Collector, collector["id"])
            assert stale_collector is not None
            stale_collector.online_status = "online"
            stale_collector.last_heartbeat_at = (datetime.now(timezone.utc) - timedelta(seconds=61)).isoformat()
            stale_collector.status_payload = {"runtime_status": "listening", "adapter_status": {}}
            db.commit()

        collector_status = client.get("/api/v1/collector-control/status", headers=headers)
        assert collector_status.status_code == 200
        stale_collector_response = next(
            item for item in collector_status.json()["collectors"] if item["id"] == collector["id"]
        )
        assert stale_collector_response["online_status"] == "offline"
        assert stale_collector_response["status_payload"]["runtime_status"] == "stale"
        assert stale_collector_response["status_payload"]["heartbeat_timeout_seconds"] == 60

        cleanup_registration = client.post(
            "/api/v1/collector-control/register",
            headers=headers,
            json={
                "collector_id": "old-simulator",
                "collector_name": "Old simulator",
                "source_machine": "old-machine",
                "client_version": "0.1.0",
            },
        )
        assert cleanup_registration.status_code == 201
        cleanup_token = cleanup_registration.json()["collector_token"]
        cleanup_collector = cleanup_registration.json()["collector"]

        cleanup_heartbeat = client.post(
            "/api/v1/collector-runtime/heartbeat",
            headers={"X-Collector-Token": cleanup_token},
            json={"source_machine": "old-machine", "adapter_status": {"simulator": {"status": "ready"}}},
        )
        assert cleanup_heartbeat.status_code == 200

        with SessionLocal() as db:
            expired_collector = db.get(Collector, cleanup_collector["id"])
            assert expired_collector is not None
            expired_collector.online_status = "online"
            expired_collector.last_heartbeat_at = (datetime.now(timezone.utc) - timedelta(hours=25)).isoformat()
            expired_collector.status_payload = {"runtime_status": "listening", "adapter_status": {}}
            db.commit()

        cleaned_status = client.get("/api/v1/collector-control/status", headers=headers)
        assert cleaned_status.status_code == 200
        assert cleanup_collector["id"] not in {item["id"] for item in cleaned_status.json()["collectors"]}

        expired_heartbeat = client.post(
            "/api/v1/collector-runtime/heartbeat",
            headers={"X-Collector-Token": cleanup_token},
            json={"source_machine": "old-machine"},
        )
        assert expired_heartbeat.status_code == 401

        with SessionLocal() as db:
            cleaned_collector = db.get(Collector, cleanup_collector["id"])
            assert cleaned_collector is not None
            assert cleaned_collector.is_deleted is True
            assert cleaned_collector.is_enabled is False
            assert cleaned_collector.token_hash is None

        upload = client.post(
            "/api/v1/collector-runtime/raw-records",
            headers={"X-Collector-Token": collector_token},
            json={
                "task_id": capture_task.json()["id"],
                "workspace_id": 999,
                "records": [
                    {
                        "document_id": "DOC-1",
                        "source_component": "simulator",
                        "source_index": "1",
                        "dedupe_key": "DOC-1-1",
                        "payload_format": "json",
                        "raw_payload": '{"order":"A001"}',
                    }
                ],
            },
        )
        assert upload.status_code == 201
        assert upload.json() == {"inserted": 1, "skipped": 0}

        duplicate_upload = client.post(
            "/api/v1/collector-runtime/raw-records",
            headers={"X-Collector-Token": collector_token},
            json={
                "task_id": capture_task.json()["id"],
                "records": [
                    {
                        "document_id": "DOC-1",
                        "source_component": "simulator",
                        "source_index": "1",
                        "dedupe_key": "DOC-1-1",
                        "payload_format": "json",
                        "raw_payload": '{"order":"A001"}',
                    }
                ],
            },
        )
        assert duplicate_upload.status_code == 201
        assert duplicate_upload.json() == {"inserted": 0, "skipped": 1}

        douyin_raw_payload = {
            "cmd": "print",
            "task": {
                "documents": [
                    {
                        "documentID": "YT123456789",
                        "contents": [
                            {
                                "data": {
                                    "trackNo": "YT123456789",
                                    "orderId": "ORDER-001",
                                    "shopName": "测试店铺",
                                    "productInfo": "【现货新款】5.0 黑白紫 38.5 1 件",
                                    "productShortInfo": "5.0 黑白紫 38.5 1 件",
                                    "productCount": "1件",
                                    "sPInfo": "SKU001 5.0 黑白紫 38.5 1 件",
                                    "templateCode": "yt_76_130_v2",
                                    "parentTemplateCode": "yt_76_130_v2",
                                }
                            }
                        ],
                    },
                    {
                        "documentID": "YT987654321",
                        "contents": [
                            {
                                "data": {
                                    "trackNo": "YT987654321",
                                    "orderId": "ORDER-002",
                                    "shopName": "测试店铺",
                                    "productInfo": "【现货新款】4.0 涂鸦 40.5 1 件",
                                    "productShortInfo": "4.0 涂鸦 40.5 1 件",
                                    "productCount": "1件",
                                    "sPInfo": "SKU002 4.0 涂鸦 40.5 1 件",
                                    "templateCode": "yt_76_130_v2",
                                    "parentTemplateCode": "yt_76_130_v2",
                                }
                            }
                        ],
                    }
                ]
            },
        }
        douyin_upload = client.post(
            "/api/v1/collector-runtime/raw-records",
            headers={"X-Collector-Token": collector_token},
            json={
                "task_id": capture_task.json()["id"],
                "records": [
                    {
                        "document_id": "YT123456789",
                        "source_component": "cloud-print-client",
                        "source_index": "2",
                        "dedupe_key": "cloud-print-client:YT123456789",
                        "payload_format": "json",
                        "raw_payload": json.dumps(douyin_raw_payload, ensure_ascii=False),
                    }
                ],
            },
        )
        assert douyin_upload.status_code == 201
        assert douyin_upload.json() == {"inserted": 1, "skipped": 0}

        cainiao_direct_payload = {
            "cmd": "print",
            "task": {
                "documents": [
                    {
                        "documentID": "YT7623405796653",
                        "contents": [
                            {
                                "encryptedData": "AES:encrypted",
                                "templateURL": "https://cloudprint.cainiao.com/template/standard/290659/94",
                            },
                            {
                                "templateURL": "https://cloudprint.cainiao.com/template/customArea/61587447/12",
                                "data": {
                                    "SHOP_NAME": "小蓝鞋铺",
                                    "ORDER_ID": "3304401529305009268",
                                    "WAIBILLNO_BAR_CODE": "YT7623405796653",
                                    "ITEM_INFO": "补.差。价。谈好多少钱。拍多少数量|提前沟通",
                                    "BUYER_MEMO": "全白史密斯44碼",
                                    "SELLER_MEMO": "注意检查瑕疵 发台湾",
                                    "PRINT_TIME": "2026-06-11 21:27:09",
                                },
                            },
                        ],
                    }
                ]
            },
        }
        cainiao_direct_upload = client.post(
            "/api/v1/collector-runtime/raw-records",
            headers={"X-Collector-Token": collector_token},
            json={
                "task_id": capture_task.json()["id"],
                "records": [
                    {
                        "document_id": "YT7623405796653",
                        "source_component": "cainiao-cnprint",
                        "source_index": "34",
                        "dedupe_key": "cainiao-direct:YT7623405796653",
                        "payload_format": "json",
                        "raw_payload": json.dumps(cainiao_direct_payload, ensure_ascii=False),
                    }
                ],
            },
        )
        assert cainiao_direct_upload.status_code == 201
        assert cainiao_direct_upload.json() == {"inserted": 1, "skipped": 0}

        cainiao_woda_payload = {
            "cmd": "print",
            "task": {
                "documents": [
                    {
                        "documentID": "605912844697866752-woda-605912837668213248+++0",
                        "contents": [
                            {
                                "encryptedData": "AES:encrypted-a",
                                "addData": {"recipient": {"mobile": "18425145758-4573"}},
                                "templateURL": "https://cloudprint.cainiao.com/template/standard/290659/94",
                            },
                            {
                                "printXML": "<text><![CDATA[【HK】特2跑步鞋，,*1]]></text>"
                                "<text><![CDATA[颜色分类:紫粉;鞋码:40]]></text>"
                            },
                        ],
                    },
                    {
                        "documentID": "605537914424662784-woda-605537907084630784+++0",
                        "contents": [
                            {
                                "encryptedData": "AES:encrypted-b",
                                "addData": {"recipient": {"mobile": "18412120587"}},
                                "templateURL": "https://cloudprint.cainiao.com/template/standard/290659/94",
                            },
                            {
                                "printXML": "<text><![CDATA[【流放】男鞋针织跑步鞋，,*1]]></text>"
                                "<text><![CDATA[颜色分类:深灰黑;尺码:43]]></text>"
                            },
                        ],
                    },
                ]
            },
        }
        cainiao_woda_upload = client.post(
            "/api/v1/collector-runtime/raw-records",
            headers={"X-Collector-Token": collector_token},
            json={
                "task_id": capture_task.json()["id"],
                "records": [
                    {
                        "document_id": "woda-batch",
                        "source_component": "cainiao-cnprint",
                        "source_index": "36",
                        "dedupe_key": "cainiao-woda:batch-36",
                        "payload_format": "json",
                        "raw_payload": json.dumps(cainiao_woda_payload, ensure_ascii=False),
                    }
                ],
            },
        )
        assert cainiao_woda_upload.status_code == 201
        assert cainiao_woda_upload.json() == {"inserted": 1, "skipped": 0}

        stop_task = client.post(
            "/api/v1/collector-control/stop",
            headers=headers,
            json={"task_id": capture_task.json()["id"]},
        )
        assert stop_task.status_code == 200
        assert stop_task.json()["status"] == "completed"

        raw_records = client.get("/api/v1/raw-capture-records", headers=headers)
        assert raw_records.status_code == 200
        simulator_record = next(item for item in raw_records.json() if item["document_id"] == "DOC-1")
        douyin_record = next(item for item in raw_records.json() if item["document_id"] == "YT123456789")
        cainiao_direct_record = next(
            item for item in raw_records.json() if item["document_id"] == "YT7623405796653"
        )
        cainiao_woda_record = next(item for item in raw_records.json() if item["document_id"] == "woda-batch")
        assert simulator_record["tenant_id"] == 1
        assert simulator_record["workspace_id"] == 1
        assert simulator_record["collector_id"] == collector["id"]
        assert simulator_record["task_id"] == capture_task.json()["id"]
        assert simulator_record["raw_payload"] == '{"order":"A001"}'
        assert douyin_record["status"] == "parsed"
        assert douyin_record["waybill_mode"] == "douyin_cloud_print"
        assert douyin_record["standard_detail_id"] is not None
        assert cainiao_direct_record["status"] == "parsed"
        assert cainiao_direct_record["waybill_mode"] == "cainiao_direct_shop"
        assert cainiao_woda_record["status"] == "parsed"
        assert cainiao_woda_record["waybill_mode"] == "cainiao_woda_printxml"
        assert cainiao_woda_record["standard_detail_id"] is not None
        assert cainiao_woda_record["parsed_payload"]["parser_status"] == "parsed"
        assert len(cainiao_woda_record["parsed_payload"]["documents"]) == 2
        assert all(
            item["usable_for_business"] is True
            for item in cainiao_woda_record["parsed_payload"]["documents"]
        )

        standard_details = client.get("/api/v1/standard-details", headers=headers)
        assert standard_details.status_code == 200
        parsed_detail = next(
            item
            for item in standard_details.json()
            if item["field_values"]["logistics_no"] == "YT123456789"
        )
        assert parsed_detail["waybill_mode"] == "douyin_cloud_print"
        assert parsed_detail["field_values"]["order_no"] == "ORDER-001"
        assert parsed_detail["field_values"]["shop_name"] == "测试店铺"
        assert parsed_detail["field_values"]["quantity"] == 1
        douyin_details = [
            item
            for item in standard_details.json()
            if item["waybill_mode"] == "douyin_cloud_print"
        ]
        assert len(douyin_details) == 2
        assert {item["field_values"]["logistics_no"] for item in douyin_details} == {
            "YT123456789",
            "YT987654321",
        }
        assert {item["field_values"]["document_sequence"] for item in douyin_details} == {1, 2}
        direct_detail = next(
            item
            for item in standard_details.json()
            if item["field_values"].get("logistics_no") == "YT7623405796653"
        )
        assert direct_detail["waybill_mode"] == "cainiao_direct_shop"
        assert direct_detail["field_values"]["source_platform"] == "cainiao_direct_shop"
        assert direct_detail["field_values"]["order_no"] == "3304401529305009268"
        assert direct_detail["field_values"]["shop_name"] == "小蓝鞋铺"
        woda_details = [
            item
            for item in standard_details.json()
            if item["waybill_mode"] == "cainiao_woda_printxml"
        ]
        assert len(woda_details) == 2
        assert {item["field_values"]["document_sequence"] for item in woda_details} == {1, 2}
        assert all(item["field_values"]["source_platform"] == "woda" for item in woda_details)
        assert all(item["field_values"]["encrypted_waybill"] is True for item in woda_details)
        assert all(item["field_values"]["print_template_source"] == "printxml_layout" for item in woda_details)
        assert len({item["field_values"]["print_template_key"] for item in woda_details}) == 1
        assert any("颜色分类:紫粉" in item["field_values"]["product_full_text"] for item in woda_details)

        woda_template_key = woda_details[0]["field_values"]["print_template_key"]
        template_config = client.post(
            "/api/v1/print-template-configs",
            headers=headers,
            json={
                "waybill_mode": "cainiao_woda_printxml",
                "template_key": woda_template_key,
                "template_label": "我打一联自定义区",
                "template_source": "printxml_layout",
                "parse_status": "custom_required",
                "config": {
                    "rule_type": "custom_text_lines_v1",
                    "product_line_index": 0,
                    "spec_line_index": 1,
                },
                "is_enabled": True,
            },
        )
        assert template_config.status_code == 201
        assert template_config.json()["tenant_id"] == 1
        assert template_config.json()["workspace_id"] == 1
        assert template_config.json()["template_key"] == woda_template_key

        force_parse = client.post(
            "/api/v1/collector-control/parse-records",
            headers=headers,
            json={"task_id": capture_task.json()["id"], "force": True},
        )
        assert force_parse.status_code == 200
        assert force_parse.json()["parsed"] == 3
        assert force_parse.json()["limited"] == 0
        reparsed_details = client.get("/api/v1/standard-details", headers=headers)
        assert reparsed_details.status_code == 200
        reparsed_woda_details = [
            item
            for item in reparsed_details.json()
            if item["waybill_mode"] == "cainiao_woda_printxml"
        ]
        assert len(reparsed_woda_details) == 2

        raw_document = client.get(
            f"/api/v1/collector-control/tasks/{capture_task.json()['id']}/raw-document",
            headers=headers,
        )
        assert raw_document.status_code == 200
        assert raw_document.content[:2] == b"PK"
        raw_workbook = load_workbook(BytesIO(raw_document.content))
        raw_sheet = raw_workbook.active
        raw_headers = [cell.value for cell in raw_sheet[1]]
        assert raw_headers == [
            "ID",
            "采集器",
            "电脑名",
            "来源组件",
            "采集时间",
            "状态",
            "采集原文",
        ]
        assert "采集任务ID" not in raw_headers
        assert "原始文档ID" not in raw_headers
        assert "来源序号" not in raw_headers
        assert "解析模式" not in raw_headers
        assert "原文格式" not in raw_headers
        assert "本地来源信息" not in raw_headers
        assert "解析诊断" not in raw_headers
        assert raw_sheet.max_row >= 5

        for export_order, (name, code) in enumerate(
            [
                ("商品简称", "product_display_text"),
                ("规格", "spec_text"),
                ("尺码", "inferred_size"),
                ("数量", "quantity"),
            ],
            start=1,
        ):
            field_definition = client.post(
                "/api/v1/field-definitions",
                headers=headers,
                json={
                    "name": name,
                    "code": code,
                    "data_type": "text",
                    "export_enabled": True,
                    "export_order": export_order,
                },
            )
            assert field_definition.status_code == 201

        standard_document = client.get(
            f"/api/v1/collector-control/tasks/{capture_task.json()['id']}/standard-document",
            headers=headers,
        )
        assert standard_document.status_code == 200
        assert standard_document.content[:2] == b"PK"
        standard_workbook = load_workbook(BytesIO(standard_document.content))
        standard_sheet = standard_workbook.active
        assert [standard_sheet.cell(row=1, column=index).value for index in range(1, 5)] == [
            "商品简称",
            "规格",
            "尺码",
            "数量",
        ]
        exported_rows = list(standard_sheet.iter_rows(min_row=2, values_only=True))
        assert any(row[0] == "5.0 黑白紫 38.5 1 件" and row[2] == "38.5" for row in exported_rows)

        rejected_upload = client.post(
            "/api/v1/collector-runtime/raw-records",
            headers={"X-Collector-Token": "bad-token"},
            json={"task_id": capture_task.json()["id"], "records": []},
        )
        assert rejected_upload.status_code == 401


def test_platform_customer_account_maintenance_flow() -> None:
    with TestClient(app) as client:
        login = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
        assert login.status_code == 200
        token = login.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}", "X-Workspace-Id": "1"}

        before = client.get("/api/v1/platform/customer-accounts", headers=headers)
        assert before.status_code == 200
        assert before.json()["workspaces"]

        created = client.post(
            "/api/v1/platform/customer-accounts",
            headers=headers,
            json={
                "tenant_name": "Customer A",
                "tenant_code": "customer_a",
                "workspace_name": "Customer A workspace",
                "workspace_code": "customer_a_ws",
                "username": "customer_admin",
                "display_name": "Customer admin",
                "password": "customer123",
            },
        )
        assert created.status_code == 201
        body = created.json()
        assert body["tenant"]["code"] == "customer_a"
        assert body["workspace"]["code"] == "customer_a_ws"
        assert body["user"]["username"] == "customer_admin"
        assert body["workspace"]["admin_users"][0]["role_name"] == "workspace_admin"

        duplicate = client.post(
            "/api/v1/platform/customer-accounts",
            headers=headers,
            json={
                "tenant_name": "Customer A",
                "tenant_code": "customer_a",
                "workspace_name": "Customer A workspace",
                "workspace_code": "customer_a_ws_2",
                "username": "customer_admin_2",
                "display_name": "Customer admin 2",
                "password": "customer123",
            },
        )
        assert duplicate.status_code == 409

        customer_login = client.post(
            "/api/v1/auth/login",
            json={"username": "customer_admin", "password": "customer123"},
        )
        assert customer_login.status_code == 200
        customer_token = customer_login.json()["access_token"]
        customer_me = client.get(
            "/api/v1/auth/me",
            headers={"Authorization": f"Bearer {customer_token}"},
        )
        assert customer_me.status_code == 200
        assert customer_me.json()["roles"] == ["workspace_admin"]
        assert customer_me.json()["workspaces"][0]["code"] == "customer_a_ws"

        reset = client.post(
            f"/api/v1/platform/customer-accounts/users/{body['user']['id']}/reset-password",
            headers=headers,
            json={"password": "customer456"},
        )
        assert reset.status_code == 200

        old_password = client.post(
            "/api/v1/auth/login",
            json={"username": "customer_admin", "password": "customer123"},
        )
        assert old_password.status_code == 401

        new_password = client.post(
            "/api/v1/auth/login",
            json={"username": "customer_admin", "password": "customer456"},
        )
        assert new_password.status_code == 200


def test_export_field_definition_upsert_restores_soft_deleted_field() -> None:
    with TestClient(app) as client:
        login = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
        assert login.status_code == 200
        token = login.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}", "X-Workspace-Id": "1"}

        created = client.post(
            "/api/v1/field-definitions",
            headers=headers,
            json={
                "name": "旧表头",
                "code": "restore_test_product_full_text",
                "data_type": "text",
                "export_enabled": True,
                "export_order": 10,
            },
        )
        assert created.status_code == 201

        deleted = client.delete(f"/api/v1/field-definitions/{created.json()['id']}", headers=headers)
        assert deleted.status_code == 204

        restored = client.post(
            "/api/v1/export-field-definitions/upsert",
            headers=headers,
            json={
                "name": "商品标题",
                "code": "restore_test_product_full_text",
                "export_order": 20,
            },
        )
        assert restored.status_code == 200
        assert restored.json()["id"] == created.json()["id"]
        assert restored.json()["name"] == "商品标题"
        assert restored.json()["export_enabled"] is True
        assert restored.json()["export_order"] == 20
        assert restored.json()["is_deleted"] is False

        records = client.get("/api/v1/field-definitions", headers=headers)
        assert any(record["code"] == "restore_test_product_full_text" for record in records.json())


def test_product_sku_zip_upload_flow() -> None:
    with TestClient(app) as client:
        login = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
        assert login.status_code == 200
        token = login.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}", "X-Workspace-Id": "1"}

        product = client.post(
            "/api/v1/products",
            headers=headers,
            json={"name": "匡威商品", "keywords": ["不应写入"], "is_enabled": True},
        )
        assert product.status_code == 201
        assert product.json()["keywords"] is None

        invalid_sku = client.post(
            "/api/v1/product-skus",
            headers=headers,
            json={"product_id": 999999, "name": "黑色"},
        )
        assert invalid_sku.status_code == 422

        buffer = BytesIO()
        with zipfile.ZipFile(buffer, "w") as archive:
            archive.writestr("SKU图/SKU图_01_匡威-黑色.png", b"image-a")
            archive.writestr("详情图/详情图_01.png", b"image-b")
        buffer.seek(0)

        upload = client.post(
            f"/api/v1/products/{product.json()['id']}/sku-zip",
            headers=headers,
            files={"file": ("sku.zip", buffer.getvalue(), "application/zip")},
        )
        assert upload.status_code == 201
        assert upload.json()["imported"] == 1
        assert upload.json()["skipped"] == 1
        assert upload.json()["skus"][0]["name"] == "黑色"
        assert upload.json()["skus"][0]["keywords"] is None

        skus = client.get("/api/v1/product-skus", headers=headers)
        assert skus.status_code == 200
        imported_sku = next(item for item in skus.json() if item["name"] == "黑色")
        assert imported_sku["product_id"] == product.json()["id"]
        assert imported_sku["image_asset_id"] is not None

        deleted_sku = client.delete(f"/api/v1/product-skus/{imported_sku['id']}", headers=headers)
        assert deleted_sku.status_code == 204

        restore_buffer = BytesIO()
        with zipfile.ZipFile(restore_buffer, "w") as archive:
            archive.writestr("SKU图/SKU图_01_匡威-黑色.png", b"image-restored")
        restore_buffer.seek(0)
        restored_upload = client.post(
            f"/api/v1/products/{product.json()['id']}/sku-zip",
            headers=headers,
            files={"file": ("sku.zip", restore_buffer.getvalue(), "application/zip")},
        )
        assert restored_upload.status_code == 201
        assert restored_upload.json()["imported"] == 1
        assert restored_upload.json()["skus"][0]["id"] == imported_sku["id"]

        skus_after_restore = client.get("/api/v1/product-skus", headers=headers)
        restored_sku = next(item for item in skus_after_restore.json() if item["name"] == "黑色")
        assert restored_sku["id"] == imported_sku["id"]
        assert restored_sku["is_deleted"] is False

        images = client.get("/api/v1/image-assets", headers=headers)
        assert images.status_code == 200
        assert any(item["name"] == "匡威商品-黑色" for item in images.json())

        image_content = client.get(
            f"/api/v1/image-assets/{restored_sku['image_asset_id']}/content",
            headers=headers,
        )
        assert image_content.status_code == 200
        assert image_content.content == b"image-restored"

        manual_upload = client.post(
            f"/api/v1/products/{product.json()['id']}/sku-image",
            headers=headers,
            data={"sku_name": "红色"},
            files={"file": ("red.webp", b"image-c", "image/webp")},
        )
        assert manual_upload.status_code == 201
        assert manual_upload.json()["sku"]["name"] == "红色"
        assert manual_upload.json()["sku"]["image_asset_id"] is not None

        skus_after_manual = client.get("/api/v1/product-skus", headers=headers)
        assert skus_after_manual.status_code == 200
        manual_sku = next(item for item in skus_after_manual.json() if item["name"] == "红色")
        assert manual_sku["product_id"] == product.json()["id"]

        manual_image_content = client.get(
            f"/api/v1/image-assets/{manual_sku['image_asset_id']}/content",
            headers=headers,
        )
        assert manual_image_content.status_code == 200
        assert manual_image_content.content == b"image-c"


def test_product_create_restores_deleted_same_name() -> None:
    with TestClient(app) as client:
        login = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
        assert login.status_code == 200
        token = login.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}", "X-Workspace-Id": "1"}

        created = client.post(
            "/api/v1/products",
            headers=headers,
            json={"name": "可恢复商品", "remark": "旧备注", "is_enabled": True},
        )
        assert created.status_code == 201

        deleted = client.delete(f"/api/v1/products/{created.json()['id']}", headers=headers)
        assert deleted.status_code == 204

        restored = client.post(
            "/api/v1/products",
            headers=headers,
            json={"name": "可恢复商品", "remark": "新备注", "is_enabled": True},
        )
        assert restored.status_code == 201
        assert restored.json()["id"] == created.json()["id"]
        assert restored.json()["is_deleted"] is False
        assert restored.json()["remark"] == "新备注"


def test_match_rule_product_target_must_match_product_id() -> None:
    with TestClient(app) as client:
        login = client.post("/api/v1/auth/login", json={"username": "admin", "password": "admin123"})
        assert login.status_code == 200
        token = login.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}", "X-Workspace-Id": "1"}

        vap = client.post("/api/v1/products", headers=headers, json={"name": "测试VAP"})
        fan_vap = client.post("/api/v1/products", headers=headers, json={"name": "测试范5 VAP"})
        assert vap.status_code == 201
        assert fan_vap.status_code == 201

        key_set = client.post(
            "/api/v1/key-field-sets",
            headers=headers,
            json={
                "name": "测试商品识别字段",
                "purpose": "product_identify",
                "field_codes": ["product_short_text"],
                "priority": 10,
                "is_enabled": True,
            },
        )
        assert key_set.status_code == 201

        valid_rule = client.post(
            "/api/v1/match-rules",
            headers=headers,
            json={
                "key_field_set_id": key_set.json()["id"],
                "match_values": {
                    "mode_code": "douyin_cloud_print",
                    "product_id": vap.json()["id"],
                    "product_name": "前端传错也会被后端覆盖",
                    "keyword": "VAP",
                    "fields": ["product_short_text"],
                },
                "target_type": "product",
                "target_id": vap.json()["id"],
                "target_name": "前端传错也会被后端覆盖",
                "priority": 100,
                "is_enabled": True,
            },
        )
        assert valid_rule.status_code == 201
        assert valid_rule.json()["target_id"] == vap.json()["id"]
        assert valid_rule.json()["target_name"] == "测试VAP"
        assert valid_rule.json()["match_values"]["product_id"] == vap.json()["id"]
        assert valid_rule.json()["match_values"]["product_name"] == "测试VAP"

        mismatched_create = client.post(
            "/api/v1/match-rules",
            headers=headers,
            json={
                "key_field_set_id": key_set.json()["id"],
                "match_values": {
                    "mode_code": "douyin_cloud_print",
                    "product_id": vap.json()["id"],
                    "keyword": "VAP",
                    "fields": ["product_short_text"],
                },
                "target_type": "product",
                "target_id": fan_vap.json()["id"],
                "priority": 100,
                "is_enabled": True,
            },
        )
        assert mismatched_create.status_code == 422

        mismatched_update = client.patch(
            f"/api/v1/match-rules/{valid_rule.json()['id']}",
            headers=headers,
            json={"target_id": fan_vap.json()["id"]},
        )
        assert mismatched_update.status_code == 422


def teardown_module() -> None:
    from app.core.database import engine

    engine.dispose()
    if TEST_DB.exists():
        TEST_DB.unlink()
    if TEST_STORAGE.exists():
        import shutil

        shutil.rmtree(TEST_STORAGE)
