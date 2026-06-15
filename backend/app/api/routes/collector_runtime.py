from datetime import datetime, timedelta, timezone
from io import BytesIO
import json
from pathlib import Path
import re
from types import SimpleNamespace
from urllib.parse import quote
from typing import Annotated, Any
from zipfile import ZIP_DEFLATED, ZipFile

from fastapi import APIRouter, Body, Depends, Header, HTTPException, Query, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from PIL import Image as PillowImage, UnidentifiedImageError
from pydantic import BaseModel, Field
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.config import get_settings
from app.core.context import CurrentUser
from app.core.database import get_db
from app.core.deps import get_current_user, get_workspace_id, require_write
from app.core.security import create_collector_token, hash_collector_token
from app.models import (
    CaptureTask,
    Collector,
    FieldDefinition,
    ImageAsset,
    MatchRule,
    Product,
    ProductSku,
    PrintTemplateConfig,
    RawCaptureRecord,
    Stall,
    StandardDetail,
    Workspace,
)
from app.repositories.base import model_to_dict
from app.services.product_recognition import int_value, recognize_detail_items, recognition_summary, text_value
from app.services.waybill_parser import parse_raw_capture_record, parse_raw_capture_records


router = APIRouter()

COLLECTOR_HEARTBEAT_TIMEOUT = timedelta(seconds=60)
COLLECTOR_CLEANUP_TIMEOUT = timedelta(hours=24)

COLLECTOR_CLIENT_ARCHIVE_ROOT = "订单系统采集器"
COLLECTOR_CLIENT_RELEASE_EXE = Path("dist") / "订单系统采集器.exe"
COLLECTOR_CLIENT_PACKAGE_VERSION = "single-exe-token-collector-20260614"


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def parse_utc_datetime(value: str | None) -> datetime | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
    except ValueError:
        return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=timezone.utc)
    return parsed.astimezone(timezone.utc)


def collector_heartbeat_is_stale(collector: Collector) -> bool:
    if collector.online_status != "online":
        return False
    last_heartbeat_at = parse_utc_datetime(collector.last_heartbeat_at)
    if last_heartbeat_at is None:
        return True
    return datetime.now(timezone.utc) - last_heartbeat_at > COLLECTOR_HEARTBEAT_TIMEOUT


def collector_should_be_cleaned(collector: Collector) -> bool:
    if collector.online_status != "online":
        return False
    last_heartbeat_at = parse_utc_datetime(collector.last_heartbeat_at)
    if last_heartbeat_at is None:
        return True
    return datetime.now(timezone.utc) - last_heartbeat_at > COLLECTOR_CLEANUP_TIMEOUT


def cleanup_collector(collector: Collector, *, user_id: int | None = None) -> None:
    status_payload = collector.status_payload if isinstance(collector.status_payload, dict) else {}
    collector.status_payload = {
        **status_payload,
        "runtime_status": "cleaned",
        "stale_reason": "heartbeat_cleanup",
        "heartbeat_cleanup_hours": int(COLLECTOR_CLEANUP_TIMEOUT.total_seconds() // 3600),
        "cleaned_at": utc_now(),
    }
    collector.online_status = "offline"
    collector.is_enabled = False
    collector.token_hash = None
    collector.is_deleted = True
    collector.updated_by = user_id


def cleanup_expired_collectors(
    db: Session,
    *,
    workspace_id: int | None = None,
    user_id: int | None = None,
) -> int:
    statement = select(Collector).where(
        Collector.is_deleted.is_(False),
        Collector.is_enabled.is_(True),
        Collector.online_status == "online",
    )
    if workspace_id is not None:
        statement = statement.where(Collector.workspace_id == workspace_id)

    cleaned_count = 0
    for collector in db.scalars(statement).all():
        if collector_should_be_cleaned(collector):
            cleanup_collector(collector, user_id=user_id)
            cleaned_count += 1
    if cleaned_count:
        db.commit()
    return cleaned_count


def get_workspace_tenant_id(db: Session, workspace_id: int) -> int | None:
    workspace = db.get(Workspace, workspace_id)
    if workspace is None or workspace.is_deleted:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Workspace access denied.")
    return workspace.tenant_id


def collector_client_source_dir() -> Path:
    return Path(__file__).resolve().parents[4] / "collector-client"


def collector_client_archive_path(name: str) -> str:
    return str(Path(COLLECTOR_CLIENT_ARCHIVE_ROOT) / name)


def write_collector_client_version(zip_file: ZipFile, *, mode: str) -> None:
    zip_file.writestr(
        collector_client_archive_path("VERSION.txt"),
        (
            f"version={COLLECTOR_CLIENT_PACKAGE_VERSION}\n"
            f"mode={mode}\n"
            "package=single-exe-token-collector\n"
            "features=single-exe,no-console-window,token-only,no-password-on-business-machine,server-reconnect-wait,remote-disconnect-guard\n"
        ),
    )


def collector_client_parameter_guide() -> str:
    return (
        "订单系统采集器参数说明\n"
        "\n"
        "文件：订单系统采集器.exe\n"
        "\n"
        "复制前只需要替换两处：\n"
        "1. <TOKEN> 换成网页后台生成的采集器 token。\n"
        "2. <服务器地址> 换成系统访问地址，例如 http://服务器IP:5173；不要填写 8000 端口。\n"
        "\n"
        "常用启动方式：\n"
        "\n"
        "1. 正式后台监听（最常用）\n"
        "订单系统采集器.exe --base-url \"<服务器地址>\" --token \"<TOKEN>\" --loop\n"
        "\n"
        "2. 指定后台显示名称（设备标识仍自动使用业务机机器名）\n"
        "订单系统采集器.exe --base-url \"<服务器地址>\" --token \"<TOKEN>\" --collector-name \"订单系统采集器\" --loop\n"
        "\n"
        "3. 先保存配置，再后台启动（后续启动命令最短）\n"
        "订单系统采集器.exe --base-url \"<服务器地址>\" --token \"<TOKEN>\" --collector-name \"订单系统采集器\" --save-config\n"
        "订单系统采集器.exe --loop\n"
        "\n"
        "4. 指定日志文件位置\n"
        "订单系统采集器.exe --base-url \"<服务器地址>\" --token \"<TOKEN>\" --loop --log-file \"%LOCALAPPDATA%\\OrderSystemCollector\\collector.log\"\n"
        "\n"
        "5. 只检查连接和本机打印组件，不持续监听\n"
        "订单系统采集器.exe --base-url \"<服务器地址>\" --token \"<TOKEN>\" --check --log-file \"%LOCALAPPDATA%\\OrderSystemCollector\\collector-check.log\"\n"
        "\n"
        "常用参数：\n"
        "--base-url        系统访问地址；不要填写 8000 端口。例如 http://服务器IP:5173。\n"
        "--token           后台生成的采集器 token，必填。业务机不再输入系统账号密码。\n"
        "--loop            持续后台监听；服务器断开或重启时不会退出，会继续等待恢复。\n"
        "--collector-name  后台显示名称，默认是 订单系统采集器。\n"
        "--interval        心跳和采集轮询间隔，默认 3 秒。\n"
        "--config          可选配置文件路径，默认保存在当前 Windows 用户的 LocalAppData。\n"
        "--state           可选状态文件路径，默认和配置文件同目录。\n"
        "--log-file        可选日志文件路径，默认和配置文件同目录 collector.log。\n"
        "--save-config     保存当前 base-url/token/名称等配置后退出；以后可直接用 订单系统采集器.exe --loop。\n"
        "--check           检查本机打印组件和服务器心跳后退出；不进入持续监听。\n"
        "\n"
        "设备标识说明：\n"
        "用户不需要填写设备标识。采集器会自动读取业务机 Windows 机器名作为设备标识并上传。\n"
        "\n"
        "后台运行说明：\n"
        "这个 exe 是无控制台窗口版本，按参数启动后不会弹黑框。日志默认写入：\n"
        "%LOCALAPPDATA%\\OrderSystemCollector\\collector.log\n"
        "\n"
        "不要使用的旧方式：\n"
        "不要在业务机运行 Python，不要用 bat/vbs，不要输入系统账号密码，不要填写后端 8000 端口。\n"
        "\n"
        "token 失效时：\n"
        "在系统后台移除旧采集器并重新生成 token，再用同一条启动命令替换 token。不要在业务机保存或输入系统登录密码。\n"
    )


def write_collector_client_release(zip_file: ZipFile, source_dir: Path, *, mode: str) -> None:
    exe_path = source_dir / COLLECTOR_CLIENT_RELEASE_EXE
    if not exe_path.is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="订单系统采集器.exe not found. Build the collector client first.",
        )
    write_collector_client_version(zip_file, mode=mode)
    zip_file.write(exe_path, collector_client_archive_path("订单系统采集器.exe"))
    zip_file.writestr(
        collector_client_archive_path("参数说明.txt"),
        collector_client_parameter_guide(),
    )


def build_collector_client_archive(mode: str = "cli") -> BytesIO:
    source_dir = collector_client_source_dir()
    if not source_dir.exists():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Collector client package not found.")

    archive = BytesIO()
    with ZipFile(archive, "w", ZIP_DEFLATED) as zip_file:
        write_collector_client_release(zip_file, source_dir, mode=mode)
    archive.seek(0)
    return archive


class CollectorRegisterRequest(BaseModel):
    collector_id: str | None = Field(default=None, max_length=128)
    collector_name: str = Field(default="订单系统采集器", max_length=128)
    source_machine: str | None = Field(default=None, max_length=128)
    client_version: str | None = Field(default=None, max_length=64)
    remark: str | None = None


class CaptureStartRequest(BaseModel):
    name: str | None = Field(default=None, max_length=128)
    collector_id: int | None = None


class CaptureStopRequest(BaseModel):
    task_id: int | None = None


class CollectorHeartbeatRequest(BaseModel):
    collector_id: str | None = Field(default=None, max_length=128)
    source_machine: str | None = Field(default=None, max_length=128)
    client_version: str | None = Field(default=None, max_length=64)
    runtime_status: str | None = Field(default=None, max_length=32)
    adapter_status: dict[str, Any] | None = None
    queue_size: int | None = None
    last_error: str | None = None


class RawCaptureRecordPayload(BaseModel):
    document_id: str | None = Field(default=None, max_length=128)
    source_machine: str | None = Field(default=None, max_length=128)
    source_component: str | None = Field(default=None, max_length=128)
    source_index: str | None = Field(default=None, max_length=128)
    dedupe_key: str | None = Field(default=None, max_length=255)
    waybill_mode: str | None = Field(default=None, max_length=128)
    payload_format: str = Field(default="unknown", max_length=32)
    raw_payload: str
    source_columns: dict[str, Any] | None = None
    parsed_payload: dict[str, Any] | None = None
    captured_at: str | None = Field(default=None, max_length=64)


class RawCaptureBatchRequest(BaseModel):
    task_id: int
    records: list[RawCaptureRecordPayload]


class ParseRecordsRequest(BaseModel):
    task_id: int | None = None
    force: bool = False


def public_collector(collector: Collector) -> dict[str, Any]:
    data = model_to_dict(collector)
    if not collector_heartbeat_is_stale(collector):
        return data

    status_payload = data.get("status_payload")
    if isinstance(status_payload, str):
        try:
            status_payload = json.loads(status_payload)
        except json.JSONDecodeError:
            status_payload = {}
    elif isinstance(status_payload, dict):
        status_payload = dict(status_payload)
    else:
        status_payload = {}

    status_payload["runtime_status"] = "stale"
    status_payload["stale_reason"] = "heartbeat_timeout"
    status_payload["heartbeat_timeout_seconds"] = int(COLLECTOR_HEARTBEAT_TIMEOUT.total_seconds())
    data["online_status"] = "offline"
    data["status_payload"] = status_payload
    return data


def public_task(task: CaptureTask) -> dict[str, Any]:
    return model_to_dict(task)


def json_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    return json.dumps(value, ensure_ascii=False, indent=2)


def text_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def source_component_label(component: Any) -> str:
    component_text = text_value(component)
    if component_text == "cloud-print-client":
        return "抖店打印组件"
    if component_text == "cainiao-cnprint":
        return "菜鸟打印组件"
    return component_text or "-"


def raw_record_collector_label(record: RawCaptureRecord, collectors_by_id: dict[int, Collector]) -> str:
    collector = collectors_by_id.get(int(record.collector_id or 0))
    if collector is not None and text_value(collector.collector_name):
        return text_value(collector.collector_name)
    return text_value(record.source_machine) or "-"


def infer_size_text(*values: Any) -> str:
    text = "\n".join(text_value(value) for value in values if text_value(value))
    labeled = re.search(r"(?:鞋码|尺码|码数|尺碼)\s*[:：]?\s*([2-4]\d(?:\.5)?|50|[XSML]{1,4})", text, re.I)
    if labeled:
        return labeled.group(1)
    generic = re.search(r"(?<!\d)([2-4]\d(?:\.5)?|50)(?!\d)", text)
    return generic.group(1) if generic else ""


CUSTOM_FIELD_FALLBACKS = {
    "custom_spec_text": "custom_sales_attr1_text",
    "custom_size_text": "custom_sales_attr2_text",
    "quantity": "custom_quantity_text",
}


def custom_item_export_values(
    base_values: dict[str, Any],
    item: dict[str, Any],
    *,
    item_index: int,
    item_count: int,
) -> dict[str, Any]:
    values = dict(base_values)
    remark_text = text_value(item.get("remark_text"))
    sales_attr1 = text_value(item.get("sales_attr1_text") or item.get("spec_text"))
    sales_attr2 = text_value(item.get("sales_attr2_text") or item.get("size_text"))
    quantity_text = text_value(item.get("quantity_text"))
    values.update(
        {
            "custom_item_index": item_index,
            "custom_item_count": item_count,
            "custom_item_key": f"{base_values.get('raw_record_id') or base_values.get('raw_document_id')}-{item_index}",
            "custom_product_text": text_value(item.get("product_text")),
            "custom_sales_attr1_text": sales_attr1,
            "custom_sales_attr2_text": sales_attr2,
            "custom_spec_text": text_value(item.get("spec_text")) or sales_attr1,
            "custom_size_text": sales_attr2,
            "custom_quantity_text": quantity_text,
            "custom_item_remark_text": remark_text,
            "custom_item_raw_text": text_value(item.get("raw_text")),
        }
    )
    if quantity_text:
        values["quantity"] = quantity_text
    return values


def standard_detail_export_rows(detail: StandardDetail) -> list[dict[str, Any]]:
    values = detail.field_values or {}
    custom_items = values.get("custom_items")
    if not isinstance(custom_items, list) or not custom_items:
        return [values]

    item_dicts = [item for item in custom_items if isinstance(item, dict)]
    if not item_dicts:
        return [values]

    item_count = len(item_dicts)
    return [
        custom_item_export_values(values, item, item_index=index, item_count=item_count)
        for index, item in enumerate(item_dicts, start=1)
    ]


def export_field_value(field_code: str, values: dict[str, Any]) -> Any:
    if field_code == "inferred_size":
        return infer_size_text(
            values.get("custom_sales_attr2_text"),
            values.get("custom_size_text"),
            values.get("custom_item_remark_text"),
            values.get("spec_text"),
            values.get("product_short_text"),
            values.get("product_full_text"),
            values.get("custom_area_raw_text"),
        )
    if field_code == "product_display_text":
        is_woda_custom_row = values.get("source_platform") == "woda" or values.get("custom_area_raw_text") not in (None, "")
        if not is_woda_custom_row:
            return (
                values.get("product_short_text")
                or values.get("product_full_text")
                or values.get("custom_item_raw_text")
                or values.get("custom_product_text")
                or values.get("custom_area_raw_text")
                or ""
            )
        return (
            values.get("custom_product_text")
            or values.get("product_short_text")
            or values.get("product_full_text")
            or values.get("custom_area_raw_text")
            or ""
        )
    value = values.get(field_code)
    if value in (None, "") and field_code in CUSTOM_FIELD_FALLBACKS:
        return values.get(CUSTOM_FIELD_FALLBACKS[field_code], "")
    return value


RECOGNITION_REPORT_HEADERS = ["商品名称", "销售属性1", "SKU图片", "销售属性2", "数量"]

RECOGNITION_REPORT_FIELD_DEFINITIONS: dict[str, dict[str, Any]] = {
    "product_name": {"label": "商品名称", "width": 16},
    "stall_name": {"label": "档口", "width": 14},
    "sales_attr1": {"label": "销售属性1", "width": 24},
    "sku_image": {"label": "SKU图片", "width": 18},
    "sales_attr2": {"label": "销售属性2", "width": 18},
    "quantity": {"label": "数量", "width": 12},
}

RECOGNITION_REPORT_DEFAULT_FIELD_ORDER = [
    "product_name",
    "sales_attr1",
    "sku_image",
    "sales_attr2",
    "quantity",
]

RECOGNITION_REPORT_OUTPUT_MODES = {"merged_sheet", "stall_sheet", "stall_workbooks"}
DEFAULT_RECOGNITION_REPORT_OUTPUT_MODE = "stall_sheet"

RECOGNITION_EXCEPTION_HEADERS = [
    "面单",
    "商品文字",
    "销售属性1",
    "销售属性2",
    "数量",
    "匹配状态",
    "异常原因",
    "备注字段",
]

REPORT_IMAGE_SIZE = 88
REPORT_ROW_HEIGHT = 86
REPORT_HEADER_ROW_HEIGHT = 26


def bounded_int(value: Any, default: int, min_value: int, max_value: int) -> int:
    try:
        parsed = int(round(float(value)))
    except (TypeError, ValueError):
        return default
    return min(max(parsed, min_value), max_value)


def default_recognition_report_layout() -> dict[str, Any]:
    return {
        "columns": [
            {
                "key": key,
                "label": RECOGNITION_REPORT_FIELD_DEFINITIONS[key]["label"],
                "visible": True,
                "width": RECOGNITION_REPORT_FIELD_DEFINITIONS[key]["width"],
            }
            for key in RECOGNITION_REPORT_DEFAULT_FIELD_ORDER
        ],
        "header_row_height": REPORT_HEADER_ROW_HEIGHT,
        "row_height": REPORT_ROW_HEIGHT,
        "image_width": REPORT_IMAGE_SIZE,
        "image_height": REPORT_IMAGE_SIZE,
        "stack_sales_attr1": False,
        "stack_sales_attr2": False,
        "output_mode": DEFAULT_RECOGNITION_REPORT_OUTPUT_MODE,
    }


def normalize_recognition_report_layout(raw_layout: Any | None = None) -> dict[str, Any]:
    default_layout = default_recognition_report_layout()
    payload = raw_layout if isinstance(raw_layout, dict) else {}
    source_columns = payload.get("columns")
    if not isinstance(source_columns, list):
        source_columns = []

    columns: list[dict[str, Any]] = []
    used_keys: set[str] = set()
    for source_column in source_columns:
        if not isinstance(source_column, dict):
            continue
        key = str(source_column.get("key") or "")
        definition = RECOGNITION_REPORT_FIELD_DEFINITIONS.get(key)
        if definition is None or key in used_keys:
            continue
        used_keys.add(key)
        label = str(source_column.get("label") or definition["label"]).strip() or definition["label"]
        columns.append(
            {
                "key": key,
                "label": label[:40],
                "visible": source_column.get("visible") is not False,
                "width": bounded_int(source_column.get("width"), int(definition["width"]), 8, 60),
            }
        )

    for key in RECOGNITION_REPORT_DEFAULT_FIELD_ORDER:
        if key in used_keys:
            continue
        definition = RECOGNITION_REPORT_FIELD_DEFINITIONS[key]
        columns.append(
            {
                "key": key,
                "label": definition["label"],
                "visible": True,
                "width": definition["width"],
            }
        )

    if not any(column["visible"] for column in columns):
        for column in columns:
            column["visible"] = True

    output_mode = str(payload.get("output_mode", payload.get("outputMode")) or default_layout["output_mode"])
    if output_mode not in RECOGNITION_REPORT_OUTPUT_MODES:
        output_mode = str(default_layout["output_mode"])

    return {
        "columns": columns,
        "header_row_height": bounded_int(
            payload.get("header_row_height", payload.get("headerRowHeight")),
            int(default_layout["header_row_height"]),
            18,
            80,
        ),
        "row_height": bounded_int(
            payload.get("row_height", payload.get("rowHeight")),
            int(default_layout["row_height"]),
            24,
            180,
        ),
        "image_width": bounded_int(
            payload.get("image_width", payload.get("imageWidth")),
            int(default_layout["image_width"]),
            32,
            220,
        ),
        "image_height": bounded_int(
            payload.get("image_height", payload.get("imageHeight")),
            int(default_layout["image_height"]),
            32,
            220,
        ),
        "stack_sales_attr1": bool(payload.get("stack_sales_attr1", payload.get("stackSalesAttr1", False))),
        "stack_sales_attr2": bool(payload.get("stack_sales_attr2", payload.get("stackSalesAttr2", False))),
        "output_mode": output_mode,
    }


def recognition_report_layout_from_query(layout: str | None) -> dict[str, Any]:
    if not layout:
        return normalize_recognition_report_layout()
    try:
        parsed = json.loads(layout)
    except json.JSONDecodeError:
        return normalize_recognition_report_layout()
    return normalize_recognition_report_layout(parsed)


def visible_recognition_report_columns(layout: dict[str, Any]) -> list[dict[str, Any]]:
    return [column for column in layout["columns"] if column.get("visible") is not False]


def recognition_status_label(status_text: str) -> str:
    return {
        "matched": "已匹配",
        "product_unmatched": "商品未命中",
        "sku_unmatched": "SKU未命中",
        "conflict": "冲突",
    }.get(status_text, status_text or "-")


def recognition_image_label(row: dict[str, Any]) -> str:
    return ""


def recognition_stall_name(row: dict[str, Any]) -> str:
    return text_value(row.get("stall_name")) or "未设置档口"


def report_quantity_value(value: Any) -> int:
    text = text_value(value)
    if not text:
        return 1
    match = re.search(r"\d+", text)
    if not match:
        return 1
    parsed = int(match.group(0))
    return parsed if parsed > 0 else 1


def strip_product_prefix(text: Any, product_name: Any) -> str:
    value = text_value(text)
    prefix = text_value(product_name)
    if prefix and value.startswith(prefix):
        return value[len(prefix):].lstrip(" -_/，,：:|")
    return value


def report_spec_text(row: dict[str, Any]) -> str:
    product_name = row.get("product_name") or ""
    for value in (row.get("sku_name"), row.get("sales_attr1_text"), row.get("product_text")):
        text = strip_product_prefix(value, product_name)
        if text:
            return text
    return "-"


def report_size_tokens(value: Any) -> list[str]:
    text = text_value(value)
    if not text:
        return ["-"]
    parts = [part for part in re.split(r"[\s,，/、]+", text) if part]
    return parts or [text]


def natural_report_sort_key(value: Any) -> tuple[int, float | str, str]:
    text = text_value(value) or "-"
    match = re.search(r"\d+(?:\.\d+)?", text)
    if match:
        return (0, float(match.group(0)), text.lower())
    return (1, text.lower(), text.lower())


def sorted_report_values(values: list[str]) -> list[str]:
    return sorted([value for value in values if value], key=natural_report_sort_key)


def expanded_sales_attr2_values(row: dict[str, Any]) -> list[str]:
    tokens = report_size_tokens(row.get("sales_attr2_text"))
    quantity = report_quantity_value(row.get("quantity_text"))
    if len(tokens) > 1:
        return tokens
    return [tokens[0] or "-"] * quantity


def recognition_report_line_items(
    rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
) -> list[dict[str, Any]]:
    normalized_layout = normalize_recognition_report_layout(layout)
    grouped: dict[tuple[Any, ...], dict[str, Any]] = {}
    ordered_keys: list[tuple[Any, ...]] = []

    for row in rows:
        if row.get("status") != "matched":
            continue
        product_category = text_value(row.get("product_name") or row.get("product_text")) or "-"
        stall_name = recognition_stall_name(row)
        spec = report_spec_text(row)
        key = (
            int_value(row.get("stall_id")) or stall_name,
            int_value(row.get("product_id")) or product_category,
            int_value(row.get("sku_id")) or spec,
            int_value(row.get("sku_image_asset_id")) or 0,
            "grouped"
            if normalized_layout["stack_sales_attr1"]
            else row.get("candidate_key")
            or f"{row.get('detail_id')}:{row.get('item_index') or 0}:{spec}:{row.get('sales_attr2_text')}",
        )
        if key not in grouped:
            grouped[key] = {
                "product_category": product_category,
                "stall_id": int_value(row.get("stall_id")),
                "stall_name": stall_name,
                "spec": spec,
                "image_label": recognition_image_label(row),
                "sku_id": int_value(row.get("sku_id")),
                "sku_image_asset_id": int_value(row.get("sku_image_asset_id")),
                "sales_attr1_values": [],
                "sales_attr2_values": [],
                "quantity": 0,
            }
            ordered_keys.append(key)

        line = grouped[key]
        line["sales_attr1_values"].append(spec)
        line["sales_attr2_values"].extend(expanded_sales_attr2_values(row))
        line["quantity"] += report_quantity_value(row.get("quantity_text"))

    report_rows: list[dict[str, Any]] = []
    for key in ordered_keys:
        row = dict(grouped[key])
        sales_attr1_values = row.pop("sales_attr1_values")
        sales_attr2_values = row.pop("sales_attr2_values")
        if normalized_layout["stack_sales_attr1"]:
            row["spec"] = " ".join(dict.fromkeys(sorted_report_values(sales_attr1_values))) or "-"
        if normalized_layout["stack_sales_attr2"]:
            row["size_text"] = " ".join(dict.fromkeys(sorted_report_values(sales_attr2_values))) or "-"
        else:
            row["size_text"] = " ".join(sorted_report_values(sales_attr2_values)) or "-"
        report_rows.append(row)

    return sorted(
        report_rows,
        key=lambda row: (
            natural_report_sort_key(row.get("stall_name")),
            natural_report_sort_key(row.get("product_category")),
            natural_report_sort_key(row.get("spec")),
            natural_report_sort_key(row.get("size_text")),
        ),
    )


def recognition_report_cell_value(row: dict[str, Any], field_key: str) -> Any:
    if field_key == "product_name":
        return row["product_category"]
    if field_key == "stall_name":
        return row["stall_name"]
    if field_key == "sales_attr1":
        return row["spec"]
    if field_key == "sku_image":
        return row["image_label"]
    if field_key == "sales_attr2":
        return row["size_text"]
    if field_key == "quantity":
        return row["quantity"]
    return ""


def recognition_report_headers(layout: dict[str, Any] | None = None) -> list[str]:
    normalized_layout = normalize_recognition_report_layout(layout)
    return [str(column["label"]) for column in visible_recognition_report_columns(normalized_layout)]


def recognition_report_export_rows(
    rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
) -> list[list[Any]]:
    normalized_layout = normalize_recognition_report_layout(layout)
    columns = visible_recognition_report_columns(normalized_layout)
    return recognition_report_export_rows_from_line_items(
        recognition_report_line_items(rows, normalized_layout),
        normalized_layout,
    )


def recognition_report_export_rows_from_line_items(
    report_rows: list[dict[str, Any]],
    layout: dict[str, Any] | None = None,
) -> list[list[Any]]:
    normalized_layout = normalize_recognition_report_layout(layout)
    columns = visible_recognition_report_columns(normalized_layout)
    return [
        [
            recognition_report_cell_value(row, str(column["key"]))
            for column in columns
        ]
        for row in report_rows
    ]


def recognition_report_rows_by_stall(report_rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in report_rows:
        grouped.setdefault(recognition_stall_name(row), []).append(row)
    if grouped:
        return grouped
    return {"未设置档口": []}


def recognition_exception_export_rows(rows: list[dict[str, Any]]) -> list[list[Any]]:
    return [
        [
            row.get("source_label") or "",
            row.get("product_text") or "",
            row.get("sales_attr1_text") or "",
            row.get("sales_attr2_text") or "",
            row.get("quantity_text") or "",
            recognition_status_label(str(row.get("status") or "")),
            row.get("reason") or "",
            row.get("remark_text") or "",
        ]
        for row in rows
        if row.get("status") != "matched"
    ]


def recognition_report_image_path(image: ImageAsset) -> Path | None:
    storage_root = Path(get_settings().storage_root).resolve()
    image_path = Path(image.file_path).resolve()
    if not image_path.is_relative_to(storage_root) or not image_path.is_file():
        return None
    return image_path


def recognition_report_image_buffer(
    image_path: Path,
    *,
    image_width: int = REPORT_IMAGE_SIZE,
    image_height: int = REPORT_IMAGE_SIZE,
) -> BytesIO | None:
    try:
        with PillowImage.open(image_path) as source:
            source.thumbnail((image_width, image_height))
            converted = source.convert("RGB")
            buffer = BytesIO()
            converted.save(buffer, format="PNG")
            buffer.seek(0)
            return buffer
    except (OSError, UnidentifiedImageError):
        return None


def attach_recognition_report_images(
    sheet,
    rows: list[dict[str, Any]],
    images_by_id: dict[int, ImageAsset],
    image_buffers: list[BytesIO],
    layout: dict[str, Any] | None = None,
) -> None:
    if not rows:
        return
    normalized_layout = normalize_recognition_report_layout(layout)
    image_column_index = next(
        (
            index
            for index, column in enumerate(visible_recognition_report_columns(normalized_layout), start=1)
            if column["key"] == "sku_image"
        ),
        None,
    )
    if image_column_index is None:
        return
    image_column_letter = get_column_letter(image_column_index)
    image_width = int(normalized_layout["image_width"])
    image_height = int(normalized_layout["image_height"])
    for row_number, row in enumerate(rows, start=2):
        image_asset_id = int_value(row.get("sku_image_asset_id"))
        if image_asset_id is None:
            continue
        image = images_by_id.get(image_asset_id)
        if image is None:
            continue
        image_path = recognition_report_image_path(image)
        if image_path is None:
            continue
        buffer = recognition_report_image_buffer(
            image_path,
            image_width=image_width,
            image_height=image_height,
        )
        if buffer is None:
            continue
        image_buffers.append(buffer)
        worksheet_image = WorksheetImage(buffer)
        worksheet_image.width = image_width
        worksheet_image.height = image_height
        sheet.add_image(worksheet_image, f"{image_column_letter}{row_number}")


def recognition_report_image_assets(
    db: Session,
    *,
    workspace_id: int,
    rows: list[dict[str, Any]],
) -> dict[int, ImageAsset]:
    image_asset_ids = sorted(
        {
            image_asset_id
            for row in rows
            if (image_asset_id := int_value(row.get("sku_image_asset_id"))) is not None
        }
    )
    if not image_asset_ids:
        return {}
    return {
        image.id: image
        for image in db.scalars(
            select(ImageAsset).where(
                ImageAsset.workspace_id == workspace_id,
                ImageAsset.id.in_(image_asset_ids),
                ImageAsset.is_deleted.is_(False),
            )
        ).all()
    }


def print_template_source_key(config: PrintTemplateConfig) -> str:
    payload = config.config if isinstance(config.config, dict) else {}
    match = payload.get("template_match")
    match_payload = match if isinstance(match, dict) else {}
    return (
        text_value(match_payload.get("source_template_key"))
        or text_value(payload.get("source_template_key"))
        or text_value(config.template_key)
    )


def enriched_recognition_rules(
    rules: list[MatchRule],
    configs: list[PrintTemplateConfig],
) -> list[Any]:
    configs_by_id = {config.id: config for config in configs}
    configs_by_key = {text_value(config.template_key): config for config in configs if text_value(config.template_key)}
    enriched_rules: list[Any] = []

    for rule in rules:
        match_values = dict(rule.match_values) if isinstance(rule.match_values, dict) else {}
        config = configs_by_id.get(int_value(match_values.get("print_template_config_id")) or 0)
        if config is None:
            config = configs_by_key.get(text_value(match_values.get("print_template_key")))

        if config is not None:
            source_key = print_template_source_key(config)
            if source_key and not text_value(match_values.get("print_template_source_key")):
                match_values["print_template_source_key"] = source_key
            if config.template_label and not text_value(match_values.get("print_template_label")):
                match_values["print_template_label"] = config.template_label

        enriched_rules.append(
            SimpleNamespace(
                id=rule.id,
                priority=rule.priority,
                is_enabled=rule.is_enabled,
                is_deleted=rule.is_deleted,
                target_type=rule.target_type,
                target_id=rule.target_id,
                target_name=rule.target_name,
                match_values=match_values,
            )
        )

    return enriched_rules


def recognition_rows_for_task(db: Session, *, workspace_id: int, task_id: int) -> list[dict[str, Any]]:
    details = standard_details_for_task(db, workspace_id=workspace_id, task_id=task_id)
    rules = db.scalars(
        select(MatchRule)
        .where(
            MatchRule.workspace_id == workspace_id,
            MatchRule.is_enabled.is_(True),
            MatchRule.is_deleted.is_(False),
        )
        .order_by(MatchRule.priority.asc(), MatchRule.id.asc())
    ).all()
    print_template_configs = db.scalars(
        select(PrintTemplateConfig).where(
            PrintTemplateConfig.workspace_id == workspace_id,
        )
    ).all()
    products = db.scalars(
        select(Product).where(
            Product.workspace_id == workspace_id,
            Product.is_enabled.is_(True),
            Product.is_deleted.is_(False),
        )
    ).all()
    skus = db.scalars(
        select(ProductSku).where(
            ProductSku.workspace_id == workspace_id,
            ProductSku.is_enabled.is_(True),
            ProductSku.is_deleted.is_(False),
        )
    ).all()
    stalls = db.scalars(
        select(Stall).where(
            Stall.workspace_id == workspace_id,
            Stall.is_enabled.is_(True),
            Stall.is_deleted.is_(False),
        )
    ).all()
    stalls_by_id = {int(stall.id): stall.name for stall in stalls}
    rows = recognize_detail_items(details, enriched_recognition_rules(rules, print_template_configs), products, skus)
    for row in rows:
        stall_id = int_value(row.get("stall_id"))
        if stall_id is not None:
            row["stall_name"] = text_value(stalls_by_id.get(stall_id))
        elif row.get("status") == "matched":
            row["stall_name"] = "未设置档口"
    relabel_recognition_rows_for_task(rows, details)
    return rows


def relabel_recognition_rows_for_task(rows: list[dict[str, Any]], details: list[StandardDetail]) -> None:
    detail_numbers = {detail.id: index for index, detail in enumerate(details, start=1)}
    for row in rows:
        detail_number = detail_numbers.get(int_value(row.get("detail_id")) or 0)
        if detail_number is None:
            continue
        item_index = int_value(row.get("item_index"))
        item_count = int_value(row.get("item_count")) or 1
        row["source_label"] = (
            f"面单 {detail_number}-{item_index}"
            if item_index and item_count > 1
            else f"面单 {detail_number}"
        )


def task_or_404(db: Session, task_id: int, workspace_id: int) -> CaptureTask:
    task = db.scalars(
        select(CaptureTask).where(
            CaptureTask.id == task_id,
            CaptureTask.workspace_id == workspace_id,
            CaptureTask.is_deleted.is_(False),
        )
    ).first()
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Capture task not found.")
    return task


def xlsx_response(workbook: Workbook, filename: str) -> StreamingResponse:
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    quoted_filename = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}; filename*=UTF-8''{quoted_filename}",
        },
    )


def zip_stream_response(buffer: BytesIO, filename: str) -> StreamingResponse:
    buffer.seek(0)
    quoted_filename = quote(filename)
    return StreamingResponse(
        buffer,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quoted_filename}"},
    )


def append_xlsx_rows(sheet, headers: list[str], rows: list[list[Any]]) -> None:
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        column_letter = column_cells[0].column_letter
        max_length = max(len(str(cell.value or "")) for cell in column_cells[:80])
        sheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)


def style_recognition_report_sheet(sheet, layout: dict[str, Any] | None = None) -> None:
    normalized_layout = normalize_recognition_report_layout(layout)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for column_index, column in enumerate(visible_recognition_report_columns(normalized_layout), start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = int(column["width"])

    sheet.row_dimensions[1].height = int(normalized_layout["header_row_height"])

    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = center
            cell.border = thin_border
            if cell.row == 1:
                cell.fill = header_fill
                cell.font = header_font
            elif not cell.value:
                cell.value = None

    for row_number in range(2, sheet.max_row + 1):
        sheet.row_dimensions[row_number].height = int(normalized_layout["row_height"])


def safe_excel_sheet_title(value: str, used_titles: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?/\\]", "_", text_value(value) or "未设置档口").strip("' ") or "未设置档口"
    title = base[:31]
    suffix = 2
    while title in used_titles:
        marker = f"_{suffix}"
        title = f"{base[:31 - len(marker)]}{marker}"
        suffix += 1
    used_titles.add(title)
    return title


def safe_download_name_part(value: str) -> str:
    return re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", text_value(value) or "未设置档口").strip() or "未设置档口"


def append_recognition_report_sheet(
    workbook: Workbook,
    *,
    title: str,
    report_rows: list[dict[str, Any]],
    report_layout: dict[str, Any],
    images_by_id: dict[int, ImageAsset],
    image_buffers: list[BytesIO],
    used_titles: set[str],
) -> None:
    sheet = workbook.create_sheet(safe_excel_sheet_title(title, used_titles))
    append_xlsx_rows(
        sheet,
        recognition_report_headers(report_layout),
        recognition_report_export_rows_from_line_items(report_rows, report_layout),
    )
    style_recognition_report_sheet(sheet, report_layout)
    attach_recognition_report_images(sheet, report_rows, images_by_id, image_buffers, report_layout)


def recognition_report_workbook(
    *,
    report_rows: list[dict[str, Any]],
    report_layout: dict[str, Any],
    images_by_id: dict[int, ImageAsset],
    sheet_title: str = "报货表",
) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    image_buffers: list[BytesIO] = []
    append_recognition_report_sheet(
        workbook,
        title=sheet_title,
        report_rows=report_rows,
        report_layout=report_layout,
        images_by_id=images_by_id,
        image_buffers=image_buffers,
        used_titles=set(),
    )
    workbook._recognition_image_buffers = image_buffers  # type: ignore[attr-defined]
    return workbook


def standard_details_for_task(db: Session, *, workspace_id: int, task_id: int) -> list[StandardDetail]:
    return [
        detail
        for detail in db.scalars(
            select(StandardDetail)
            .where(
                StandardDetail.workspace_id == workspace_id,
                StandardDetail.is_deleted.is_(False),
            )
            .order_by(StandardDetail.id.asc())
        ).all()
        if int((detail.field_values or {}).get("capture_task_id") or 0) == task_id
    ]


def get_collector_from_token(
    db: Session,
    x_collector_token: str | None,
) -> Collector:
    if not x_collector_token:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Missing collector token.")

    token_hash = hash_collector_token(x_collector_token)
    collector = db.scalars(
        select(Collector).where(
            Collector.token_hash == token_hash,
            Collector.is_enabled.is_(True),
            Collector.is_deleted.is_(False),
        )
    ).first()
    if collector is None:
        raise HTTPException(status_code=status.HTTP_401_UNAUTHORIZED, detail="Invalid collector token.")
    if collector_should_be_cleaned(collector):
        cleanup_collector(collector)
        db.commit()
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Collector heartbeat expired. Please register again.",
        )
    return collector


def active_task_statement(workspace_id: int, collector_db_id: int | None = None):
    statement = select(CaptureTask).where(
        CaptureTask.workspace_id == workspace_id,
        CaptureTask.status == "collecting",
        CaptureTask.is_deleted.is_(False),
    )
    if collector_db_id is not None:
        statement = statement.where(
            (CaptureTask.collector_id.is_(None)) | (CaptureTask.collector_id == collector_db_id)
        )
    return statement.order_by(CaptureTask.id.desc())


def upsert_collector(
    db: Session,
    *,
    tenant_id: int | None,
    workspace_id: int,
    payload: CollectorRegisterRequest,
    user_id: int | None,
) -> tuple[Collector, str]:
    token = create_collector_token()
    token_hash = hash_collector_token(token)
    collector_identity = payload.collector_id or f"collector-{token[:12]}"

    collector = db.scalars(
        select(Collector).where(
            Collector.workspace_id == workspace_id,
            Collector.collector_id == collector_identity,
            Collector.is_deleted.is_(False),
        )
    ).first()
    if collector is None:
        collector = Collector(
            tenant_id=tenant_id,
            workspace_id=workspace_id,
            collector_id=collector_identity,
            collector_name=payload.collector_name,
            token_hash=token_hash,
            source_machine=payload.source_machine,
            client_version=payload.client_version,
            online_status="offline",
            remark=payload.remark,
            created_by=user_id,
            updated_by=user_id,
        )
        db.add(collector)
    else:
        collector.collector_name = payload.collector_name
        collector.token_hash = token_hash
        collector.source_machine = payload.source_machine
        collector.client_version = payload.client_version
        collector.is_enabled = True
        collector.remark = payload.remark
        collector.updated_by = user_id

    db.commit()
    db.refresh(collector)
    return collector, token


def collector_identity_is_available(
    db: Session,
    *,
    workspace_id: int,
    collector_identity: str,
    current_collector_id: int,
) -> bool:
    existing = db.scalars(
        select(Collector).where(
            Collector.workspace_id == workspace_id,
            Collector.collector_id == collector_identity,
            Collector.id != current_collector_id,
            Collector.is_deleted.is_(False),
        )
    ).first()
    return existing is None


@router.get("/collector-client/download")
def download_collector_client(
    mode: str = Query(default="cli", pattern="^(cli|script|exe)$"),
    _current_user: CurrentUser = Depends(get_current_user),
) -> StreamingResponse:
    archive = build_collector_client_archive(mode)
    filename = quote("订单整理系统采集器.zip")
    return StreamingResponse(
        archive,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{filename}"},
    )


@router.post("/collector-control/register", status_code=status.HTTP_201_CREATED)
def register_collector(
    payload: CollectorRegisterRequest = Body(default_factory=CollectorRegisterRequest),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    tenant_id = get_workspace_tenant_id(db, workspace_id)
    collector, token = upsert_collector(
        db,
        tenant_id=tenant_id,
        workspace_id=workspace_id,
        payload=payload,
        user_id=current_user.id,
    )
    return {"collector": public_collector(collector), "collector_token": token}


@router.get("/collector-control/status")
def collector_status(
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    cleanup_expired_collectors(db, workspace_id=workspace_id, user_id=current_user.id)
    collectors = db.scalars(
        select(Collector)
        .where(Collector.workspace_id == workspace_id, Collector.is_deleted.is_(False))
        .order_by(Collector.id.desc())
    ).all()
    active_task = db.scalars(active_task_statement(workspace_id)).first()
    return {
        "collectors": [public_collector(collector) for collector in collectors],
        "active_task": public_task(active_task) if active_task else None,
    }


@router.post("/collector-control/start", status_code=status.HTTP_201_CREATED)
def start_capture(
    payload: CaptureStartRequest = Body(default_factory=CaptureStartRequest),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    existing = db.scalars(active_task_statement(workspace_id)).first()
    if existing is not None:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail="A capture task is already collecting.")

    tenant_id = get_workspace_tenant_id(db, workspace_id)
    if payload.collector_id is not None:
        collector = db.get(Collector, payload.collector_id)
        if collector is None or collector.workspace_id != workspace_id or collector.is_deleted:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Collector access denied.")

    task = CaptureTask(
        tenant_id=tenant_id,
        workspace_id=workspace_id,
        name=payload.name or f"采集任务 {utc_now()}",
        collector_id=payload.collector_id,
        status="collecting",
        started_at=utc_now(),
        config={"started_by": current_user.id},
        created_by=current_user.id,
        updated_by=current_user.id,
    )
    db.add(task)
    db.commit()
    db.refresh(task)
    return public_task(task)


@router.post("/collector-control/stop")
def stop_capture(
    payload: CaptureStopRequest = Body(default_factory=CaptureStopRequest),
    db: Session = Depends(get_db),
    current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    if payload.task_id is None:
        task = db.scalars(active_task_statement(workspace_id)).first()
    else:
        task = db.scalars(
            select(CaptureTask).where(
                CaptureTask.id == payload.task_id,
                CaptureTask.workspace_id == workspace_id,
                CaptureTask.is_deleted.is_(False),
            )
        ).first()
    if task is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Active capture task not found.")
    if task.status != "collecting":
        return public_task(task)

    config = dict(task.config or {})
    config["ended_by"] = current_user.id
    task.config = config
    task.status = "completed"
    task.ended_at = utc_now()
    task.updated_by = current_user.id
    db.commit()
    db.refresh(task)
    return public_task(task)


@router.post("/collector-control/parse-records")
def parse_capture_records(
    payload: ParseRecordsRequest = Body(default_factory=ParseRecordsRequest),
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(require_write),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, int]:
    statement = select(RawCaptureRecord).where(
        RawCaptureRecord.workspace_id == workspace_id,
        RawCaptureRecord.is_deleted.is_(False),
    )
    if payload.task_id is not None:
        statement = statement.where(RawCaptureRecord.task_id == payload.task_id)
    records = db.scalars(statement.order_by(RawCaptureRecord.id.asc())).all()
    result = parse_raw_capture_records(db, records, force=payload.force)
    db.commit()
    return result


@router.get("/collector-control/tasks/{task_id}/raw-document")
def download_raw_capture_document(
    task_id: int,
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> StreamingResponse:
    task = task_or_404(db, task_id, workspace_id)
    records = db.scalars(
        select(RawCaptureRecord)
        .where(
            RawCaptureRecord.workspace_id == workspace_id,
            RawCaptureRecord.task_id == task.id,
            RawCaptureRecord.is_deleted.is_(False),
        )
        .order_by(RawCaptureRecord.id.asc())
    ).all()
    collector_ids = sorted({int(record.collector_id) for record in records if record.collector_id})
    collectors_by_id = {
        collector.id: collector
        for collector in db.scalars(
            select(Collector).where(
                Collector.id.in_(collector_ids),
                Collector.workspace_id == workspace_id,
                Collector.is_deleted.is_(False),
            )
        ).all()
    } if collector_ids else {}

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "原文"
    rows = [
        [
            record.id,
            raw_record_collector_label(record, collectors_by_id),
            record.source_machine,
            source_component_label(record.source_component),
            record.captured_at,
            record.status,
            record.raw_payload,
        ]
        for record in records
    ]
    append_xlsx_rows(
        sheet,
        [
            "ID",
            "采集器",
            "电脑名",
            "来源组件",
            "采集时间",
            "状态",
            "采集原文",
        ],
        rows,
    )
    return xlsx_response(workbook, f"capture-task-{task.id}-raw.xlsx")


@router.get("/collector-control/tasks/{task_id}/standard-document")
def download_standard_capture_document(
    task_id: int,
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> StreamingResponse:
    task = task_or_404(db, task_id, workspace_id)
    details = standard_details_for_task(db, workspace_id=workspace_id, task_id=task.id)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "整理结果"
    export_fields = db.scalars(
        select(FieldDefinition)
        .where(
            FieldDefinition.workspace_id == workspace_id,
            FieldDefinition.export_enabled.is_(True),
            FieldDefinition.export_order > 0,
            FieldDefinition.is_deleted.is_(False),
        )
        .order_by(FieldDefinition.export_order.asc(), FieldDefinition.id.asc())
    ).all()

    if not export_fields:
        append_xlsx_rows(
            sheet,
            ["提示"],
            [["当前工作区还没有定义整理文档表头，暂不生成业务整理文档。"]],
        )
        return xlsx_response(workbook, f"capture-task-{task.id}-standard.xlsx")

    rows = []
    for detail in details:
        for values in standard_detail_export_rows(detail):
            rows.append(
                [export_field_value(field.code, values) for field in export_fields]
            )
    append_xlsx_rows(
        sheet,
        [field.name for field in export_fields],
        rows,
    )
    return xlsx_response(workbook, f"capture-task-{task.id}-standard.xlsx")


@router.get("/collector-control/tasks/{task_id}/recognition-preview")
def preview_capture_task_recognition(
    task_id: int,
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> dict[str, Any]:
    task = task_or_404(db, task_id, workspace_id)
    details = standard_details_for_task(db, workspace_id=workspace_id, task_id=task.id)
    rows = recognition_rows_for_task(db, workspace_id=workspace_id, task_id=task.id)
    return {
        "task_id": task.id,
        "task_name": task.name,
        "detail_count": len(details),
        "rows": rows,
        "summary": recognition_summary(rows),
    }


@router.get("/collector-control/tasks/{task_id}/recognition-report")
def download_capture_task_recognition_report(
    task_id: int,
    layout: str | None = Query(default=None),
    db: Session = Depends(get_db),
    _current_user: CurrentUser = Depends(get_current_user),
    workspace_id: int = Depends(get_workspace_id),
) -> StreamingResponse:
    task = task_or_404(db, task_id, workspace_id)
    rows = recognition_rows_for_task(db, workspace_id=workspace_id, task_id=task.id)
    images_by_id = recognition_report_image_assets(db, workspace_id=workspace_id, rows=rows)
    report_layout = recognition_report_layout_from_query(layout)
    report_rows = recognition_report_line_items(rows, report_layout)
    exception_rows = recognition_exception_export_rows(rows)

    if report_layout["output_mode"] == "stall_workbooks":
        archive = BytesIO()
        with ZipFile(archive, "w", ZIP_DEFLATED) as zip_file:
            for stall_name, stall_rows in recognition_report_rows_by_stall(report_rows).items():
                stall_workbook = recognition_report_workbook(
                    report_rows=stall_rows,
                    report_layout=report_layout,
                    images_by_id=images_by_id,
                    sheet_title=safe_download_name_part(stall_name),
                )
                workbook_buffer = BytesIO()
                stall_workbook.save(workbook_buffer)
                zip_file.writestr(f"{safe_download_name_part(stall_name)}.xlsx", workbook_buffer.getvalue())
            if exception_rows:
                exception_workbook = Workbook()
                exception_sheet = exception_workbook.active
                exception_sheet.title = "异常明细"
                append_xlsx_rows(exception_sheet, RECOGNITION_EXCEPTION_HEADERS, exception_rows)
                exception_buffer = BytesIO()
                exception_workbook.save(exception_buffer)
                zip_file.writestr("异常明细.xlsx", exception_buffer.getvalue())
        return zip_stream_response(archive, f"capture-task-{task.id}-recognition-report-by-stall.zip")

    workbook = Workbook()
    image_buffers: list[BytesIO] = []
    if report_layout["output_mode"] == "stall_sheet":
        workbook.remove(workbook.active)
        used_titles: set[str] = set()
        for stall_name, stall_rows in recognition_report_rows_by_stall(report_rows).items():
            append_recognition_report_sheet(
                workbook,
                title=stall_name,
                report_rows=stall_rows,
                report_layout=report_layout,
                images_by_id=images_by_id,
                image_buffers=image_buffers,
                used_titles=used_titles,
            )
    else:
        sheet = workbook.active
        sheet.title = "报货表"
        append_xlsx_rows(
            sheet,
            recognition_report_headers(report_layout),
            recognition_report_export_rows_from_line_items(report_rows, report_layout),
        )
        style_recognition_report_sheet(sheet, report_layout)
        attach_recognition_report_images(sheet, report_rows, images_by_id, image_buffers, report_layout)

    if exception_rows:
        exception_sheet = workbook.create_sheet("异常明细")
        append_xlsx_rows(exception_sheet, RECOGNITION_EXCEPTION_HEADERS, exception_rows)
    return xlsx_response(workbook, f"capture-task-{task.id}-recognition-report.xlsx")


@router.post("/collector-runtime/heartbeat")
def collector_heartbeat(
    payload: CollectorHeartbeatRequest = Body(default_factory=CollectorHeartbeatRequest),
    db: Session = Depends(get_db),
    x_collector_token: Annotated[str | None, Header(alias="X-Collector-Token")] = None,
) -> dict[str, Any]:
    collector = get_collector_from_token(db, x_collector_token)
    collector.online_status = "online"
    collector.last_heartbeat_at = utc_now()
    collector.status_payload = {
        "runtime_status": payload.runtime_status or "unknown",
        "adapter_status": payload.adapter_status or {},
        "queue_size": payload.queue_size,
        "last_error": payload.last_error,
        "received_at": utc_now(),
    }
    if payload.source_machine:
        collector.source_machine = payload.source_machine
    reported_identity = str(payload.collector_id or payload.source_machine or "").strip()
    if reported_identity and collector_identity_is_available(
        db,
        workspace_id=collector.workspace_id,
        collector_identity=reported_identity,
        current_collector_id=collector.id,
    ):
        collector.collector_id = reported_identity
    if payload.client_version:
        collector.client_version = payload.client_version

    tasks = db.scalars(active_task_statement(collector.workspace_id, collector.id)).all()
    db.commit()
    return {
        "collector": public_collector(collector),
        "tasks": [public_task(task) for task in tasks],
    }


@router.post("/collector-runtime/raw-records", status_code=status.HTTP_201_CREATED)
def upload_raw_records(
    payload: RawCaptureBatchRequest,
    db: Session = Depends(get_db),
    x_collector_token: Annotated[str | None, Header(alias="X-Collector-Token")] = None,
) -> dict[str, int]:
    collector = get_collector_from_token(db, x_collector_token)
    task = db.get(CaptureTask, payload.task_id)
    if (
        task is None
        or task.workspace_id != collector.workspace_id
        or task.is_deleted
        or task.status not in {"collecting", "completed"}
        or (task.collector_id is not None and task.collector_id != collector.id)
    ):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Capture task access denied.")

    inserted = 0
    skipped = 0
    for item in payload.records:
        if item.dedupe_key:
            existing = db.scalars(
                select(RawCaptureRecord).where(
                    RawCaptureRecord.workspace_id == collector.workspace_id,
                    RawCaptureRecord.dedupe_key == item.dedupe_key,
                    RawCaptureRecord.is_deleted.is_(False),
                )
            ).first()
            if existing is not None:
                skipped += 1
                continue

        record = RawCaptureRecord(
            tenant_id=collector.tenant_id,
            workspace_id=collector.workspace_id,
            task_id=task.id,
            collector_id=collector.id,
            document_id=item.document_id,
            source_machine=item.source_machine or collector.source_machine,
            source_component=item.source_component,
            source_index=item.source_index,
            dedupe_key=item.dedupe_key,
            waybill_mode=item.waybill_mode,
            payload_format=item.payload_format,
            raw_payload=item.raw_payload,
            source_columns=item.source_columns,
            parsed_payload=item.parsed_payload,
            captured_at=item.captured_at or utc_now(),
            status="pending",
        )
        db.add(record)
        db.flush()
        parse_raw_capture_record(db, record)
        inserted += 1

    db.commit()
    return {"inserted": inserted, "skipped": skipped}
